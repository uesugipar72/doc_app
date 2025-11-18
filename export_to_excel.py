import os
import sys
import json
import openpyxl
from openpyxl.styles import Font, Alignment

def export_to_excel(data, headers,  output_folder="export_folder", filename="filename.xlsx"):
    # 出力フォルダが存在しなければ作成
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, filename)


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "検索結果"

    # ヘッダーの書き込み
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # データ行の書き込み
    for row_num, row_data in enumerate(data, start=2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # カラム幅自動調整
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    wb.save(output_file)
    print(f"Excelファイル '{output_file}' を保存しました。")

if __name__ == "__main__":
    # コマンドライン引数からデータとヘッダーを取得
    if len(sys.argv) != 5:
        print("使用法: python export_to_excel.py '<データJSON>' '<ヘッダーJSON>'")
        sys.exit(1)

    json_data = sys.argv[1]
    json_headers = sys.argv[2]
    output_folder = sys.argv[3]
    filename = sys.argv[4]

    data = json.loads(json_data)
    headers = json.loads(json_headers)

    export_to_excel(data, headers, output_folder, filename)